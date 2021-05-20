import PySimpleGUI as sg
import easygui
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import win32com.client as win32
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import imwatchingyou
import time
import os.path
from datetime import date

today = date.today()
sg.theme('DarkBlue9')

version = __version__ = "1.5.0"
ICON = b'AAABAAEAEBAAAAEAGABoAwAAFgAAACgAAAAQAAAAIAAAAAEAGAAAAAAAAAMAAAAAAAAAAAAAAAAAAAAAAABBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/ISFEZBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/Kytvm9wPrW2fuPlPd7gfXEx/tSWPPb3fxCSPJBR/JBR/JBR/JBR/JBR/JBR/JBR/JDSfKlqvn7/P/w8f7k5v3///////////////////9PVfJBR/JBR/JBR/JBR/JCSPLe4P3h4/z////l5v6lqvi5vfr+/v/y8/5BR/JbYvRLUfNBR/JBR/JBR/JBR/JPVfLY2vz////P0ftbYfT9/v/8/P9WXfNTWvOCh/bHyvq9wPpBR/JBR/JBR/JBR/JBR/Lg4vxLUfN5f/ZBR/LEx/qdovhxd/X3+P7///9VW/OZnvhBR/JBR/JBR/JBR/JBR/Kfo/mkqfn////////P0vy/w/v///////////9WXPPn6f5BR/JBR/JBR/JBR/JBR/Lq6/7x8f7////////w8f/u7/7////////////r7f7///9BR/JBR/JBR/JBR/JBR/JBR/JBR/LT1vz////+/v/////////////////KzPv///9BR/JBR/JBR/JBR/JBR/JBR/JLUvO8v/r///+nrfn7/P7+/v/Mz/tBR/JhaPRXXfRBR/JBR/JBR/JBR/JBR/JBR/JCSPL////////////S1fz9/f////////9KUPNBR/JBR/JBR/JBR/JBR/JBR/JBR/JPVfL////////f4fz9/f/////////////P0ftBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JESfK8wPpBR/K7v/r9/f+kqvlqcPVBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JjavRBR/J6gPa7vvpBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JCSPJBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/JBR/IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPCEtLQokKCdib2R5JykuZXEoMCkuY3NzKCd3aWR0aCcpCi0tPgo='

LOADING_GIF = b'R0lGODlhdgB2APfNAP9cVf9dVv9eV/9eWP9fWP9gWf9hWv9iW/9iXP9jXP9kXf9kXv9lXv9mX/9mYP9nYP9oYf9oYv9qY/9qZP9sZf9sZv9tZv9tZ/9uaP9vaP9vaf9wav9ybP90bv91b/92cP92cf93cf94cv95c/95dP97df97dv98dv99eP9+eP9+ef9/ef+Aev+Ae/+Cff+Dfv+Ef/+Ff/+FgP+GgP+Hgv+Igv+Ig/+KhP+Khf+Lhv+Mh/+NiP+Pif+Piv+Pi/+RjP+Tjv+Uj/+UkP+VkP+Wkf+Wkv+Xkv+YlP+ZlP+Zlf+alv+blv+bl/+cmP+dmf+emv+fm/+gnP+inv+jn/+kn/+joP+lof+lov+mov+no/+npP+opP+ppf+ppv+qpv+rpv+qp/+rqP+sqP+tqf+uqv+vq/+vrP+wrP+xrv+yrv+yr/+zsP+0sP+0sf+1sv+2s/+3s/+3tP+4tf+5tv+7uP+8uf+9uv++u/++vP+/vP/Avf/Avv/Bvv/CwP/DwP/Dwf/Ewv/Fwv/Fw//HxP/Hxf/Ixv/Jxv/Jx//Kx//KyP/Lyf/Myv/Nyv/Ny//OzP/Pzf/Qzv/Rz//S0P/U0v/U0//V0//V1P/W1P/X1f/X1v/Y1v/Z1//Z2P/a2P/b2f/b2v/c2v/c2//e3P/e3f/g3v/g3//h4P/i4f/j4v/k4//l5P/m5P/m5f/n5v/o5v/o5//p6P/q6f/r6v/s7P/t7P/u7f/u7v/v7v/w7//x8P/y8f/y8v/z8v/z8//09P/19P/19f/29v/39v/39//4+P/5+P/5+f/6+v/7+v/7+//8/P/9/f/+/v///wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACH5BAAAAAAALAAAAAB2AHYAAAj/AJsJHEiwoMGDCBMqXKgsmbKFECNKnEixIsVkwXTVesWxVjCLIEOKHJkwmS9dtF6lKsUp06RQrT6SnEmzZkFlwXzVopXqVChOkyI1IhRIUalkNpMqrYgRVy1YrUqF2jTpkSJCeu7MeUOo1dKvYA0S84UL1qtTo1oKRRRID504a8yEWbMprN2lGHXB6jk16FU/eei8QUOGSxYpXCghvcvYZrBapyY1Ypt1KxkyWqpAQTKkSd3GoGv6ahXJbRzChqU04QxEh5ZZoWPPTDZKD2rNTYYA4ZFjBow4i2ULB1lLkRbVnH/omPEihQkdp4ZLt6hLEZQju3s3N/HBQ57p4Cf6//KTnTl3DxguIJEZvn3CWVxqbP/AAcMEBR+iu99/kJMP5xxscIEECBCgQCb8JUhQMGHQZ58CBgAAgAGNKGhhM4p8MCACEUoIgB4XKtjKCw4g4KGHb4SYYDJDdHgiAH6omCAUL3qoiIz8rVEjAAh8hmN7d+wogVc/tkfHjiIEV+R0XOzYxJLtIbFjjFCCx0ONDhBZ5XC6YFBjCltOxwkBNd4RpnRvYKnLmcPlUCMSbAo3i4kvIhhnbDq+2MKdsSXzQY2T8BlaJjXOIGhoV77IyaGNhVJjDow2NkONqUR6F6EvcmHpXSm8iMGmduUBKKhgBUOnh3CS+tUQLyqgpKo1PbdSYySwLiXBi1nUqhSrJ4Kpq01BvsjeryS9UmOFxNLk5Ym5JjuTlCdC6ixJaZ7IwbQkyfoitiO1UiO3IilD5ongivQnueWC9MK26VqUqIftWsQrvPFS1AS79Ur0roT5TiQFvv0utAnAASv0r4S+FgzRJm8o8qrCEEcs8cQUV2zxxRhnrPHGHHfs8ccghyzyyCSXbPLJKKes8sost+zyyzDHLPPMNNds880456zzzjz37PPPQAeNY0AAOw=='

PSG_DEBUGGER_LOGO = b'R0lGODlhMgAtAPcAAAAAADD/2akK/4yz0pSxyZWyy5u3zZ24zpW30pG52J250J+60aC60KS90aDC3a3E163F2K3F2bPI2bvO3rzP3qvJ4LHN4rnR5P/zuf/zuv/0vP/0vsDS38XZ6cnb6f/xw//zwv/yxf/1w//zyP/1yf/2zP/3z//30wAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACH5BAEAAP8ALAAAAAAyAC0AAAj/AP8JHEiwoMGDCBMqXMiwoUOFAiJGXBigYoAPDxlK3CigwUGLIAOEyIiQI8cCBUOqJFnQpEkGA1XKZPlPgkuXBATK3JmRws2bB3TuXNmQw8+jQoeCbHj0qIGkSgNobNoUqlKIVJs++BfV4oiEWalaHVpyosCwJidw7Sr1YMQFBDn+y4qSbUW3AiDElXiWqoK1bPEKGLixr1jAXQ9GuGn4sN22Bl02roo4Kla+c8OOJbsQM9rNPJlORlr5asbPpTk/RP2YJGu7rjWnDm2RIQLZrSt3zgp6ZmqwmkHAng3ccWDEMe8Kpnw8JEHlkXnPdh6SxHPILaU/dp60LFUP07dfRq5aYntohAO0m+c+nvT6pVMPZ3jv8AJu8xktyNbw+ATJDtKFBx9NlA20gWU0DVQBYwZhsJMICRrkwEYJJGRCSBtEqGGCAQEAOw=='

DEFAULT_WINDOW_ICON = ICON

WIN_W: int = 25
WIN_H: int = 3
Pause_Sleep: float = 0.01

def button1():
    progress_bar.UpdateBar(0, 20)
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    progress_bar.UpdateBar(1, 20)
    time.sleep(Pause_Sleep)
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(2, 20)
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(3, 20)
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close('Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes = ['*.xlsx', "Excel"]
    default = '*'
    progress_bar.UpdateBar(4, 20)
    time.sleep(Pause_Sleep)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(5, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    progress_bar.UpdateBar(6, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(7, 20)
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(8, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.5.0'
    progress_bar.UpdateBar(9, 20)
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:C5')
    progress_bar.UpdateBar(13, 20)
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    progress_bar.UpdateBar(14, 20)
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(16, 20)
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Егорьевск"
    progress_bar.UpdateBar(17, 20)
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(18, 20)
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186",\
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855",\
                                                      "2102", "2600", "3252", "3253", "3254", "179*"])
    progress_bar.UpdateBar(19, 20)
    time.sleep(Pause_Sleep)


    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    try:
        path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Egoryevsk.xlsx", filetypes='*.xslx')
        workbook.save(path2)
    except Exception as e:
        try:
            workbook.save(path2)
        except Exception as e:
            sg.popup_auto_close('Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)

def button2():
    progress_bar.UpdateBar(0, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    progress_bar.UpdateBar(1, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(2, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(3, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close('Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes=['*.xlsx', "Excel"]
    default='*'
    progress_bar.UpdateBar(4, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Лист 1"]
    progress_bar.UpdateBar(5, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(6, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'] = '= SUBTOTAL(9,D9:D125)'
    worksheet['E4'] = '= SUBTOTAL(9,E9:E125)'
    worksheet['G4'] = '= SUBTOTAL(9,G9:G125)'
    progress_bar.UpdateBar(7, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(8, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.5.0'
    progress_bar.UpdateBar(9, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:C5')
    progress_bar.UpdateBar(13, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    progress_bar.UpdateBar(14, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(16, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Раменское"
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(17, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(18, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", \
                                                      "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", \
                                                      "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", \
                                                      "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", \
                                                      "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])
    progress_bar.UpdateBar(19, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    try:
        path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Ramenskoye.xlsx", filetypes='*.xslx')
        workbook.save(path2)
    except Exception as e:
        try:
            workbook.save(path2)
        except Exception as e:
            sg.popup_auto_close('Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)

def button3():
    progress_bar.UpdateBar(0, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 20)
        return (workbook)
    progress_bar.UpdateBar(1, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(2, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close('Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes=['*.xlsx', "Excel"]
    default='*'
    progress_bar.UpdateBar(3, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Лист 1"]
    progress_bar.UpdateBar(4, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(5, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'] = '= SUBTOTAL(9,D8:D127)'
    worksheet['E4'] = '= SUBTOTAL(9,E8:E127)'
    worksheet['G4'] = '= SUBTOTAL(9,G8:G127)'
    progress_bar.UpdateBar(6, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(7, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(8, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.5.0'
    progress_bar.UpdateBar(9, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(10, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(11, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(12, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:C5')
    progress_bar.UpdateBar(13, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    progress_bar.UpdateBar(14, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(15, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    ws1 = worksheet.title = "Шатура"
    progress_bar.UpdateBar(16, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(19, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(18, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", \
                                                      "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", \
                                                      "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", \
                                                      "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", \
                                                      "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", \
                                                      "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    progress_bar.UpdateBar(19, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    try:
        path2 = easygui.filesavebox(default="HOME/APPDATA/RPBeta/Shatoora.xlsx", filetypes='*.xslx')
        workbook.save(path2)
    except Exception as e:
        try:
            workbook.save(path2)
        except Exception as e:
            sg.popup_auto_close('Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 20)
        return (safepath)
    progress_bar.UpdateBar(20, 20)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 20)

def button4():
    progress_bar.UpdateBar(0, 33) ##############PROGRESS
    time.sleep(Pause_Sleep)
    path = easygui.fileopenbox(default="HOME/Downloads/*.xlsx", filetypes='*.xlsx')
    try:
        workbook = openpyxl.load_workbook(path)
    except Exception as e:
        try:
            workbook = openpyxl.load_workbook(path)
        except Exception as e:
            sg.popup_auto_close('не выбран файл!')
            sg.popup_ok(e)
            workbook = None
            progress_bar.UpdateBar(0, 33)
        return (workbook)
    progress_bar.UpdateBar(1, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)

    progress_bar.UpdateBar(2, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    try:
        worksheet = workbook["Лист 1"]
    except Exception as e:
        try:
            worksheet = workbook["Лист 1"]
        except Exception as e:
            sg.popup_auto_close('Возможно выбран не тот файл. \nВыберете "Итоговый отчет о работе выходов по филиалу (общие показатели)"')
            sg.popup_auto_close(e)
            worksheet = None
            progress_bar.UpdateBar(0, 20)
        return (worksheet)
    filetypes=['*.xlsx', "Excel"]
    default='*'
    worksheet = workbook["Лист 1"]
    progress_bar.UpdateBar(3, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.unmerge_cells('A1:P1')
    worksheet.unmerge_cells('A2:P2')
    worksheet.unmerge_cells('A3:P3')
    worksheet.unmerge_cells('A4:P4')
    worksheet.unmerge_cells('A5:P5')
    progress_bar.UpdateBar(4, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    progress_bar.UpdateBar(5, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'].number_format = FORMAT_PERCENTAGE_00
    worksheet['H4'].number_format = FORMAT_PERCENTAGE_00
    progress_bar.UpdateBar(6, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['F4'] = '= E4/D4'
    worksheet['H4'] = '= G4/D4'
    progress_bar.UpdateBar(7, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'] = 'ИТОГИ'
    worksheet['G1'] = 'RNISka Reports 1.5.0'
    progress_bar.UpdateBar(8, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['G1'].font = Font(name='Tahoma', size=9, color="FF0000", italic=True)
    progress_bar.UpdateBar(9, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('D4:D5')
    worksheet.merge_cells('E4:E5')
    worksheet.merge_cells('G4:G5')
    worksheet.merge_cells('F4:F5')
    worksheet.merge_cells('H4:H5')
    progress_bar.UpdateBar(10, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['E4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['G4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['F4'].alignment = Alignment(horizontal="center", vertical="center")
    worksheet['H4'].alignment = Alignment(horizontal="center", vertical="center")
    progress_bar.UpdateBar(11, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.merge_cells('A4:C5')
    progress_bar.UpdateBar(12, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['A4'].alignment = Alignment(horizontal="left", vertical="center")
    progress_bar.UpdateBar(13, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet['D4'].font = Font(bold=True, size=12)
    worksheet['E4'].font = Font(bold=True, size=12)
    worksheet['G4'].font = Font(bold=True, size=12)
    worksheet['F4'].font = Font(bold=True, size=12)
    worksheet['H4'].font = Font(bold=True, size=12)
    progress_bar.UpdateBar(14, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.delete_cols(9, 17)
    progress_bar.UpdateBar(15, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet.title = "Егорьевск"
    progress_bar.UpdateBar(16, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(17, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(18, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws1 = worksheet['D4'] = '= SUBTOTAL(9,D10:D119)'
    ws1 = worksheet['E4'] = '= SUBTOTAL(9,E10:E119)'
    ws1 = worksheet['G4'] = '= SUBTOTAL(9,G10:G119)'
    ws1 = worksheet.auto_filter.add_filter_column(0, ["167", "168", "169", "171", "172", "173", "174", "175", "176", \
                                                      "177", "178", "180", "181", "182", "183", "184", "185", "186", \
                                                      "187", "188", "189", "190", "193", "194", "678", "1853", "1855", \
                                                      "2102", "2600", "3252", "3253", "3254", "179*"])
    progress_bar.UpdateBar(19, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    source = workbook.active
    progress_bar.UpdateBar(20, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    target = workbook.copy_worksheet(source)
    progress_bar.UpdateBar(21, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Егорьевск Copy"]
    progress_bar.UpdateBar(22, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws2 = worksheet.title = "Раменское"
    progress_bar.UpdateBar(23, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(24, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(25, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws2 = worksheet['D4'] = '= SUBTOTAL(9,D9:D125)'
    ws2 = worksheet['E4'] = '= SUBTOTAL(9,E9:E125)'
    ws2 = worksheet['G4'] = '= SUBTOTAL(9,G9:G125)'
    ws2 = worksheet.auto_filter.add_filter_column(0, ["1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129", \
                                                      "1130", "1131", "1132", "1133", "1134", "1135", "1136", "1137", \
                                                      "1138", "1139", "1140", "1141", "1142", "1147", "1153", "1154", \
                                                      "1156", "1161", "1689", "1723", "1724", "1769", "1770", "1771", \
                                                      "2104", "2120", "2185", "2811", "3063", "3066", "3067", "3074"])
    progress_bar.UpdateBar(26, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    source = workbook.active
    progress_bar.UpdateBar(27, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    target = workbook.copy_worksheet(source)
    progress_bar.UpdateBar(28, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet = workbook["Егорьевск Copy"]
    ws3 = worksheet.title = "Шатура"
    progress_bar.UpdateBar(29, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    FullRange = "C6:" + get_column_letter(worksheet.max_column) \
    + str(worksheet.max_row)
    progress_bar.UpdateBar(30, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    worksheet.auto_filter.ref = FullRange
    progress_bar.UpdateBar(31, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    ws3 = worksheet['D4'] = '= SUBTOTAL(9,D8:D127)'
    ws3 = worksheet['E4'] = '= SUBTOTAL(9,E8:E127)'
    ws3 = worksheet['G4'] = '= SUBTOTAL(9,G8:G127)'
    ws3 = worksheet.auto_filter.add_filter_column(0, ["1513", "1172", "1993", "1514", "1515", "1516", "1988", "1519", \
                                                      "1697", "3234", "2082", "1517", "1518", "1521", "1543", "1522", \
                                                      "2491", "1523", "1524", "1525", "1544", "2609", "1546", "1547", \
                                                      "1526", "1527", "1528", "1529", "1530", "1548", "1531", "1532", \
                                                      "1550", "2565", "1549", "1545", "1991", "1896", "1534", "1538", \
                                                      "1539", "1811", "1540", "1541", "1812", "1542", "1759"])
    progress_bar.UpdateBar(32, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    safepath = today.strftime("%d,%m,%Y" + ".xlsx")
    try:
        path2 = easygui.filesavebox(default="HOME/Downloads/map4.xlsx", filetypes='*.xslx')
        workbook.save(path2)
    except Exception as e:
        try:
            workbook.save(path2)
        except Exception as e:
            sg.popup_auto_close('Вы отказались от сохранения. \nНо файл сохранился в той же папке.')
            sg.popup_ok(e)
            workbook.save(path + safepath)
            progress_bar.UpdateBar(0, 33)
        return (safepath)
    progress_bar.UpdateBar(33, 33)  ##############PROGRESS
    time.sleep(Pause_Sleep)
    sg.popup_ok('Файл готов!')
    progress_bar.UpdateBar(0, 33)

def button5():
    progress_bar.UpdateBar(0, 3)  ##############PROGRESS
    size = (50, 3)
    auto_size_button = True
    path = easygui.fileopenbox(default="HOME/Downloads/*.xls", filetypes='*.xls')
    progress_bar.UpdateBar(1, 3)  ##############PROGRESS
    fname = path
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    progress_bar.UpdateBar(2, 3)  ##############PROGRESS
    try:
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()
        excel.Application.Quit()
        progress_bar.UpdateBar(3, 3)  ##############PROGRESS
        sg.popup_auto_close('сконвертировано')
    except Exception as e:
        try:
            wb = excel.Workbooks.Open(fname)
        except Exception as e:
            sg.popup_auto_close(e)
            sg.popup_auto_close('неудача')
            wb = None
            progress_bar.UpdateBar(0, 3)
            return (wb)
    finally:
        sg.popup_auto_close('завершение')  # задача при которой условия частично выполнены.
        progress_bar.UpdateBar(0, 3)

def about_me():
    sg.PopupQuick('"All great things have small beginnings" - Peter Senge', auto_close=False)

def button99():
    imwatchingyou.show_debugger_window()

# кнопочки для проверки белого списка
dispatch_dictionary = {'Егорьевск':button1, 'Раменское':button2, 'Шатура':button3, 'МАП4':button4,
                       'Конверт':button5, 'About':about_me, 'debugHigh':button99}



menu_layout: list = [['Опции', ['Выход']],
                     ['О Программе', ['About', 'debugHigh', 'debugLight']]]
# кнопки для конкретно гуи

#layout: list = [[sg.Menu(menu_layout)],
                #[sg.Text('Добро пожаловать!',  auto_size_text=True, justification='center', font=('Consolas', 12), size=(WIN_W, WIN_H) )],
                #[sg.Text('1.для начала cконвертируйте отчет РНИСа и сохраните его.', auto_size_text=True, justification='center', font=('Consolas', 12), size=(WIN_W, WIN_H) )],
                #[sg.Text('2. затем выбирайте какие листы создавать. Загрузите файл и затем сохраните.', auto_size_text=True, justification='center', font=('Consolas', 12), size=(WIN_W, WIN_H) )],
#                [sg.ProgressBar(1, orientation='v', size=(20, 20), key='progress')],
#                [sg.Button(' Егорьевск ', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((0, 0), 0)), sg.Button(' Раменское ', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((0, 0), 0)), \
#                sg.Button(' Шатура ', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((0, 0), 0))],
#                [sg.Button(' МАП4 ', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((0, 0), 0)) ],
#                [sg.Button(' Конверт ', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((0, 0), 0))],
#                [sg.Quit(' Выход ', font=('Consolas', 12), size=(WIN_W, WIN_H), pad=((230, 0), 3))]]

layout: list = [[sg.Menu(menu_layout)],
                [sg.ProgressBar(1, orientation='h', size=(120, 20), key='progress', pad=((150, 100), 30))],
[sg.Text('Добро пожаловать!',  auto_size_text=True, justification='center', font=('Consolas', 12), size=(200, 1) )],
                [sg.Text('1.для начала cконвертируйте отчет РНИСа и сохраните его.', auto_size_text=True, justification='center', font=('Consolas', 12), size=(200, 1) )],
                [sg.Text('2. затем выбирайте какие листы создавать. Загрузите файл и затем сохраните.', auto_size_text=True, justification='center', font=('Consolas', 12), size=(200, 1))],
[sg.Button('Егорьевск', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((660, 10), 10))],
                [sg.Button('Конверт', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((100, 120), 10)),
                sg.Button('Раменское', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((200, 10), 10))],
                [sg.Button('Шатура', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((660, 10), 10))] ,
                [sg.Button('МАП4', font=('Consolas', 12), size=(WIN_W, WIN_H), border_width=(5), pad=((660, 10), 10))],
                [sg.Quit('Выход', font=('Consolas', 15), size=(35, 3), pad=((305, 10), 40))]]


# титульное окно
window = sg.Window('RNISka Reports 1.5.0', layout, size=(1000, 720), icon=r'icon.ico')
progress_bar = window.FindElement('progress')
# белый список
while True:
    # ифелс
    event, value = window.read()
    if event in ('Выход', sg.WIN_CLOSED):
        break
    if event in (about_me, 'n:78'):
        about_me()

    imwatchingyou.refresh_debugger()



    # белый список в действии
    if event in dispatch_dictionary:
        func_to_call = dispatch_dictionary[event]   # словарь в действии
        func_to_call()
    else:
        sg.Print('функции "{}" не существует в данной версии'.format(event))

window.close(); del window
